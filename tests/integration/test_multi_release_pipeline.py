import os
import shutil
from numbers import Real

import pytest
from openpyxl import load_workbook

from thai_data_platform.pipeline import run_pipeline

pytestmark = pytest.mark.integration


def test_new_release_is_loaded_once_and_can_be_marked_as_backfill(tmp_path):
    if os.getenv("RUN_FULL_INTEGRATION") != "1":
        pytest.skip("Set RUN_FULL_INTEGRATION=1 to run the full PostgreSQL/ClickHouse path")
    postgres_url = os.getenv("POSTGRES_URL")
    if not postgres_url:
        pytest.skip("POSTGRES_URL is not configured")

    baseline_cgd = "datasets/cgd/2026.07.03.xlsx"
    release_two = tmp_path / "cgd-release-2.xlsx"
    shutil.copy2(baseline_cgd, release_two)
    _change_allocated_measure(release_two)

    common = {
        "ocsc_path": "datasets/ocsc/thai-gov-manpower-2567.4.xlsx",
        "postgres_url": postgres_url,
        "clickhouse_host": os.getenv("CLICKHOUSE_HOST", "localhost"),
        "clickhouse_port": int(os.getenv("CLICKHOUSE_HTTP_PORT", "8123")),
        "clickhouse_user": os.getenv("CLICKHOUSE_USER", "default"),
        "clickhouse_password": os.getenv("CLICKHOUSE_PASSWORD", ""),
        "raw_root": tmp_path / "raw",
    }
    run_pipeline(cgd_path=baseline_cgd, **common)

    backfill = run_pipeline(
        cgd_path=release_two,
        run_type="backfill",
        **common,
    )

    assert backfill.status == "serving_published"
    assert backfill.run_type == "backfill"
    assert backfill.core_counts["budget_facts"] > 0
    assert backfill.serving_counts["skipped_existing_sources"] == 1


def _change_allocated_measure(path):
    workbook = load_workbook(path)
    for worksheet in workbook.worksheets:
        for header_row in range(1, min(worksheet.max_row, 8) + 1):
            for column in range(1, worksheet.max_column + 1):
                header = str(worksheet.cell(header_row, column).value or "")
                if "จัดสรร" not in header:
                    continue
                for row in range(header_row + 1, worksheet.max_row + 1):
                    cell = worksheet.cell(row, column)
                    if isinstance(cell.value, Real) and not isinstance(cell.value, bool):
                        cell.value = cell.value + 1
                        workbook.save(path)
                        return
    workbook.properties.title = "release-two-fixture"
    workbook.save(path)
