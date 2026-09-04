from datetime import date

from openpyxl import Workbook

from thai_data_platform.ingestion.metadata import SourceFileMetadata
from thai_data_platform.transform.cgd import extract_cgd_workbook
from thai_data_platform.transform.ocsc import extract_ocsc_workbook


def _metadata(path, dataset_name):
    return SourceFileMetadata.from_file(
        dataset_name=dataset_name,
        source_name="synthetic public source",
        path=path,
    )


def test_cgd_parser_keeps_report_grain_and_signed_gap(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "หน่วยงาน"
    sheet["A2"] = "ข้อมูล ณ วันที่ 3 กรกฎาคม 2569"
    sheet["B4"] = "หน่วยงาน"
    sheet["D4"] = "รวม"
    for column, label in enumerate(
        ["วงเงิน", "จัดสรร", "เบิกจ่าย", "ร้อยละเบิกจ่าย", "การใช้จ่าย", "ร้อยละใช้จ่าย"],
        start=4,
    ):
        sheet.cell(5, column, label)
    sheet["B6"] = "1. กรมตัวอย่าง"
    sheet["D6"] = 100
    sheet["E6"] = 90
    sheet["F6"] = 30
    sheet["G6"] = 30
    sheet["H6"] = 40
    sheet["I6"] = 40
    sheet["B7"] = "รวม"
    sheet["D7"] = 100
    sheet["E7"] = 90
    sheet["F7"] = 30
    sheet["G7"] = 30
    sheet["H7"] = 40
    sheet["I7"] = 40
    path = tmp_path / "cgd.xlsx"
    workbook.save(path)

    extract = extract_cgd_workbook(path, _metadata(path, "cgd_budget_execution"), "run-1")

    assert len(extract.budget_execution) == 2
    assert extract.budget_execution.iloc[0]["entity_name"] == "กรมตัวอย่าง"
    assert extract.budget_execution.iloc[1]["entity_type"] == "total"
    assert extract.budget_execution.iloc[0]["disbursement_pct"] == 30
    assert extract.as_of_date == date(2026, 7, 3)
    assert len(extract.raw_cells) > 0
    assert len(extract.workbook_sheets) == 1


def test_ocsc_parser_separates_person_and_percentage_metrics(tmp_path):
    workbook = Workbook()
    agency = workbook.active
    agency.title = "หน่วยงาน"
    agency["A1"] = "ข้าราชการ ลูกจ้างประจำ"
    agency["A2"] = "พนักงานราชการ"
    agency["A6"] = "กระทรวงตัวอย่าง"
    agency["D6"] = 100
    agency["A7"] = None
    agency["C7"] = "กรมตัวอย่าง"
    agency["D7"] = 25

    profile = workbook.create_sheet("ข้อมูลช่วงอายุ")
    profile["A1"] = "ช่วงอายุ"
    profile["B1"] = "เพศ"
    profile["C1"] = "ระดับการศึกษา"
    profile["A4"] = "กรมตัวอย่าง"
    profile["B4"] = 100
    profile["C4"] = 10
    profile["P4"] = 52.5

    path = tmp_path / "ocsc.xlsx"
    workbook.save(path)

    extract = extract_ocsc_workbook(
        path,
        _metadata(path, "ocsc_government_manpower"),
        "run-1",
    )

    assert len(extract.workforce_agency) == 2
    assert len(extract.workforce_profile) == 3
    assert set(extract.workforce_agency["source_unit"]) == {"person"}
    profile_rows = extract.workforce_profile.set_index("metric_name")
    assert profile_rows.loc["age_lt_21", "source_unit"] == "person"
    assert profile_rows.loc["female_pct", "source_unit"] == "pct"
    assert profile_rows.loc["female_pct", "percentage"] == 52.5
